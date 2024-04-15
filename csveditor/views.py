from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.http import JsonResponse
from django.core.files.base import ContentFile
from django.utils import timezone
import csv
import io
import pandas as pd
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
import zipfile
from openpyxl import load_workbook
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from .models import UserProfile
from django.contrib.auth import authenticate, login

def index(request):
    if 'email' not in request.session:
        return redirect('sign')  # Assuming your login URL name is 'login'
    return render(request, 'index.html')

def merge(request):
    if 'email' not in request.session:
        return redirect('sign')  # Assuming your login URL name is 'login'
    return render(request, 'merge.html')

def converttocsv(request):
    if 'email' not in request.session:
        return redirect('sign')  # Assuming your login URL name is 'login'
    return render(request, 'xlsx2csv.html')
def sign(request):
    return render(request, 'sign.html')


def signup(request):
    if request.method == 'POST':
        # Retrieve form data
        username = request.POST.get('username')
        number = request.POST.get('number')
        email = request.POST.get('email')
        country = request.POST.get('country')
        password = request.POST.get('password')
        confirmpassword = request.POST.get('confirmpassword')

        # Check if passwords match
        if password != confirmpassword:
            return JsonResponse({'error': 'Passwords do not match'}, status=400)

        # Process the data (you can perform further validation and save to database here)
        # For now, just return a success response
        user_profile = UserProfile.objects.create(
            username=username,
            number=number,
            email=email,
            country=country,
            password=password
        )



        return JsonResponse({'success': True})

    # If the request method is not POST, return an error response
    return JsonResponse({'error': 'Method not allowed'}, status=405)


def signin(request):
    if request.method == 'POST':
        username = request.POST.get('loginusername')
        password = request.POST.get('loginpassword')

        # Authenticate user
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            # Set session email
            request.session['email'] = user.username  # Assuming username is the email
            return JsonResponse({'message': 'Login successful'}, status=200)
        else:
            # If user authentication fails, try with UserProfile model
            try:
                user_profile = UserProfile.objects.get(username=username)
                if user_profile.password == password:
                    # Set session email
                    request.session['email'] = user_profile.email
                    print(request.session['email'])
                    return JsonResponse({'message': 'Login successful'}, status=200)
                else:
                    return JsonResponse({'error': 'Invalid username or password'}, status=400)
            except UserProfile.DoesNotExist:
                return JsonResponse({'error': 'User does not exist'}, status=400)

    # If the request method is not POST, return an error response
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@csrf_exempt
def removefiltered(request):
    if request.method == 'POST':
        # Get the row title and parameter from the form data
        row_title = request.POST.get('row_title', '')
        parameters = request.POST.get('parameter', '').split(',')  # Split parameters by comma

        # Handle the file upload
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            # Process the uploaded file
            fs = FileSystemStorage(location=settings.MEDIA_ROOT)
            saved_file = fs.save(uploaded_file.name, uploaded_file)
            file_content = uploaded_file.read().decode('utf-8')

            # Parse the CSV data
            csv_data = io.StringIO(file_content)
            reader = csv.reader(csv_data)

            # Filter the rows based on the parameters in the given row
            filtered_rows = []
            header = next(reader)
            row_title_index = header.index(row_title)
            for row in reader:
                for parameter in parameters:
                    if row[row_title_index].startswith(parameter.strip()):
                        filtered_rows.append(row)
                        break  # Break out of inner loop if matched to next parameter

            # Create a new CSV file with the filtered rows
            output_csv = io.StringIO()
            writer = csv.writer(output_csv)
            writer.writerow(header)
            writer.writerows(filtered_rows)
            output_csv.seek(0)

            # Prepare response with the new CSV file as attachment
            response = HttpResponse(output_csv.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="filtered_data.csv"'
            return response

    # If the request method is not POST or file processing fails,
    # return a method not allowed response or an error message
    return HttpResponse("Error processing file or invalid request", status=400)


@csrf_exempt
def savefiltered(request):
    if request.method == 'POST':
        # Get the row title and parameter from the form data
        row_title = request.POST.get('row_title', '')
        parameters = request.POST.get('parameter', '').split(',')  # Split parameters by comma

        # Handle the file upload
        uploaded_file = request.FILES.get('file')
        if uploaded_file:
            # Process the uploaded file
            file_content = uploaded_file.read().decode('utf-8')

            # Parse the CSV data
            csv_data = io.StringIO(file_content)
            reader = csv.reader(csv_data)

            # Filter the rows based on the parameters in the given row
            filtered_rows = []
            header = next(reader)
            row_title_index = header.index(row_title)
            for row in reader:
                if not any(row[row_title_index].startswith(parameter.strip()) for parameter in parameters):
                    filtered_rows.append(row)

            # Create a new CSV file with the filtered rows
            output_csv = io.StringIO()
            writer = csv.writer(output_csv)
            writer.writerow(header)
            writer.writerows(filtered_rows)
            output_csv.seek(0)

            # Prepare response with the new CSV file as attachment
            response = HttpResponse(output_csv.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="remaining_data.csv"'
            return response

    # If the request method is not POST or file processing fails,
    # return a method not allowed response or an error message
    return HttpResponse("Error processing file or invalid request", status=400)


@csrf_exempt
def split(request):
    if request.method == 'POST':
        # Get the uploaded file and number of rows per file from the request
        uploaded_file = request.FILES.get('file')
        num_rows_per_file = int(request.POST.get('num_rows', 0))

        if uploaded_file and num_rows_per_file > 0:
            # Process the uploaded file
            file_content = uploaded_file.read().decode('utf-8')

            # Parse the CSV data
            csv_data = io.StringIO(file_content)
            reader = csv.reader(csv_data)
            header = next(reader)

            # Split the CSV data into chunks of specified number of rows
            file_chunks = []  # Initialize the list of chunks
            current_chunk = [header]  # Initialize the current chunk with header row
            row_count = 0

            for row in reader:
                current_chunk.append(row)
                row_count += 1

                # Check if the current chunk size reaches the specified number of rows per file
                if row_count >= num_rows_per_file:
                    file_chunks.append(current_chunk)  # Append the current chunk to the list of chunks
                    current_chunk = []  # Reset the current chunk
                    row_count = 0

            # Append the remaining rows to the last chunk
            if current_chunk:
                file_chunks.append(current_chunk)

            # Create a ZIP file containing the split files
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
                for index, chunk in enumerate(file_chunks):
                    # Create a new CSV file in the ZIP archive for each chunk
                    chunk_csv = io.StringIO()
                    writer = csv.writer(chunk_csv)
                    writer.writerows(chunk)

                    # Write the CSV data to the ZIP file
                    zip_file.writestr(f'file_{index + 1}.csv', chunk_csv.getvalue())

            # Prepare response with the ZIP file as attachment
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="split_files.zip"'
            return response

    # If the request method is not POST or file processing fails,
    # return a method not allowed response or an error message
    return HttpResponse("Error processing file or invalid request", status=400)

@csrf_exempt
def mergefiles(request):
    if request.method == 'POST' and request.FILES.getlist('files'):
        # Get the list of uploaded files
        uploaded_files = request.FILES.getlist('files')
        
        # Initialize a list to store the content of each CSV file
        csv_content = []

        # Read the content of each CSV file and store it in the csv_content list
        for file in uploaded_files:
            try:
                decoded_file = file.read().decode('utf-8')
                csv_content.append(decoded_file.splitlines())
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=400)
        
        # Merge the CSV content
        merged_csv_content = []
        for content in csv_content:
            csv_reader = csv.reader(content)
            for row in csv_reader:
                merged_csv_content.append(row)

        # Return the merged CSV content as a JSON response
        return JsonResponse({'merged_csv_content': merged_csv_content})

    return JsonResponse({'error': 'No files were uploaded'}, status=400)

@csrf_exempt
def tocsv(request):
    if request.method == 'POST' and request.FILES.getlist('files'):
        # Get the list of uploaded files
        uploaded_files = request.FILES.getlist('files')
        
        # Initialize a list to store the content of each CSV file
        csv_content = []

        # Read each XLSX file and convert to CSV
        for file in uploaded_files:
            try:
                # Load the XLSX file
                wb = load_workbook(file, data_only=True)
                # Get the active sheet
                sheet = wb.active

                # Create a CSV file in memory
                csv_data = []
                for row in sheet.iter_rows(values_only=True):
                    csv_data.append(row)

                # Add the CSV data to the list
                csv_content.append(csv_data)
            except Exception as e:
                return HttpResponse(str(e), status=400)
        
        # Create a new CSV file with the merged content
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="merged_file.csv"'
        writer = csv.writer(response)
        for data in csv_content:
            writer.writerows(data)

        return response

    return HttpResponse('No files were uploaded', status=400)


def alphabet_to_index(alphabet):
    # Convert the alphabet to uppercase to handle lowercase input
    alphabet = alphabet.upper()
    # Convert the alphabet to its corresponding index number (A=1, B=2, ..., Z=26)
    index = ord(alphabet) - 64  # Subtract 64 because 'A' is ASCII 65
    if index > 9:
        index -= 1  # Adjust index for letters after 'I'
    return index

def clean_phone_number(number):
    # Remove special characters and letters
    cleaned_number = ''.join(filter(str.isdigit, number))
    
    # Check if the number starts with '0'
    if cleaned_number.startswith('0'):
        # Replace '0' with '254'
        cleaned_number = '254' + cleaned_number[1:]

    # Check if the cleaned number has at least 9 digits
    if len(cleaned_number) >= 9:
        return cleaned_number
    else:
        return None  # Return None for numbers less than 9 digits


def cleandata(request):
    if request.method == 'POST' and request.FILES.get('file') and request.POST.get('rowheader'):
        # Get the uploaded CSV file and row header from the request
        uploaded_file = request.FILES['file']
        row_header = request.POST['rowheader']

        # Convert alphabet letter to column number if necessary
        if row_header.isalpha():
            try:
                header_index = alphabet_to_index(row_header)
            except ValueError:
                return HttpResponse('Invalid column letter', status=400)
        else:
            return HttpResponse('Invalid column input', status=400)

        # Parse the CSV file and extract the data from the specified column
        cleaned_data = set()  # Use a set to store unique cleaned numbers
        try:
            reader = csv.reader(uploaded_file.read().decode('utf-8').splitlines())
            header = next(reader)  # Skip the header row
            if 0 <= header_index < len(header):
                for row in reader:
                    cleaned_number = clean_phone_number(row[header_index])  # Clean the phone number in the specified column
                    if cleaned_number:  # Check if the number is not None (i.e., has at least 9 digits)
                        cleaned_data.add(cleaned_number)  # Add the cleaned number to the set
            else:
                return HttpResponse('Invalid column number', status=400)
        except Exception as e:
            return HttpResponse('Error cleaning data: ' + str(e), status=400)

        # Create a new CSV file with the cleaned data
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="cleaned_data.csv"'
        writer = csv.writer(response)
        # writer.writerow(['Cleaned Phone Number'])  # Write the header for the cleaned phone numbers
        for cleaned_number in cleaned_data:
            writer.writerow([cleaned_number])  # Write the cleaned number to the CSV

        return response

    return HttpResponse('Invalid request method or parameters', status=400)



